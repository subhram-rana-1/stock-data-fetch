# Runs simulation from start date to end date, generates entry and exit points in momentum catching
# and stores in output xlsx sheet
import os
import openpyxl
from datetime import datetime, date, time, timedelta
from typing import List

from price_app.classes import PriceData, PriceDataPerTick
from price_app.handlers import fetch_price_data
from price_app.utils import get_occurrence_distribution
from stock_data_fetch.enums import MarketType
from price_app import configs

# CONSTANTS ---------
direction_up = 'up'
direction_down = 'down'
market_entry_time = time(9, 16)
market_exit_time = time(15, 29)
exit_reason_crossover = 'crossover'
exit_reason_sl_hit = 'sl_hit'
exit_reason_fix_target_achieved = 'fix_target_achieved'
date_str_format = "%Y-%m-%d"
time_str_format = "%H:%M:%S"


# CONFIGS ---------
class Config:
    # chart configs
    smooth_price_averaging_method = configs.smooth_price_averaging_method
    smooth_price_period = configs.smooth_price_period
    smooth_price_ema_period = configs.smooth_price_ema_period

    smooth_slope_averaging_method = configs.smooth_slope_averaging_method
    smooth_slope_period = configs.smooth_slope_period
    smooth_slope_ema_period = configs.smooth_slope_ema_period

    smooth_momentum_averaging_method = configs.smooth_momentum_averaging_method
    smooth_momentum_period = configs.smooth_momentum_period
    smooth_momentum_ema_period = configs.smooth_momentum_ema_period

    # trading configs
    min_entry_time = time(9, 20)
    entry_config_smooth_slope_threshold = 0.2
    entry_config_momentum_threshold = 0.4
    entry_config_momentum_rate_threshold = 0.1
    sl = 5
    fix_point_diff_target = 3


# INPUTS ----------
class Input:
    market_type = MarketType.NIFTY
    start_date_time = datetime(2024, 9, 23, 9, 15, 0)
    end_date_time = datetime(2024, 9, 23, 15, 30, 0)
    # end_date_time = datetime(2024, 9, 19, 15, 29, 0)

    momentum_occurrence_distribution_bucket_size = 5

    output_file_relative_path = 'price_app/scripts/momentum_analysis/momentum_analysis_output.xlsx'
    sheet_name = 'momentum'
    clean_sheet = True

    result_start_row = 2

    col_date = 1
    col_entry_time = 2
    col_exit_time = 3
    col_exp_direction = 4
    col_entry_price = 5
    col_exit_price = 6
    col_move = 7
    col_exit_reason = 8

    col_momentum_buckets = 10
    col_occurrence = 11
    col_occurrence_cum_sum = 12
    col_occurrence_sum_percentage = 13


class MarketMoveData:
    def __init__(
            self,
            date: date,
            delta: float,
            expected_direction: str,
            start_time: time,
            end_time: time,
            start_point: float,
            end_point: float,
            exit_reason: str,
    ):
        self.date: date = date
        self.delta: date = delta
        self.expected_direction: str = expected_direction
        self.start_time: time = start_time
        self.end_time: time = end_time
        self.start_point: float = start_point
        self.end_point: float = end_point
        self.exit_reason: str = exit_reason

    def to_dict(self) -> dict:
        return {
            'date': self.date.strftime(date_str_format),
            'delta': round(self.delta, 2),
            'expected_direction': self.expected_direction,
            'start_time': self.start_time.strftime(time_str_format),
            'end_time': self.end_time.strftime(time_str_format),
            'start_point': round(self.start_point, 2),
            'end_point': round(self.end_point, 2),
            'exit_reason': self.exit_reason,
        }


def entry_condition_for_up_move(
        smooth_slope: float,
        entry_config_smooth_slope_threshold: float,
        momentum: float,
        entry_config_momentum_threshold: float,
        momentum_rate: float,
        entry_config_momentum_rate_threshold: float,
) -> bool:
    return \
            smooth_slope >= entry_config_smooth_slope_threshold and \
            momentum >= entry_config_momentum_threshold and \
            momentum_rate >= entry_config_momentum_rate_threshold


def entry_condition_for_down_move(
        smooth_slope: float,
        entry_config_smooth_slope_threshold: float,
        momentum: float,
        entry_config_momentum_threshold: float,
        momentum_rate: float,
        entry_config_momentum_rate_threshold: float,
) -> bool:
    return \
            smooth_slope <= -1 * entry_config_smooth_slope_threshold and \
            momentum <= -1 * entry_config_momentum_threshold and \
            momentum_rate <= -1 * entry_config_momentum_rate_threshold


def cross_over_happened_in_up_move(cur_smooth_price: float, cur_smooth_price_ema: float) -> bool:
    return cur_smooth_price < cur_smooth_price_ema


def cross_over_happened_in_down_move(cur_smooth_price: float, cur_smooth_price_ema: float) -> bool:
    return cur_smooth_price > cur_smooth_price_ema


def trail_sl_in_upward_expected_move(cur_tick_price: float, sl_diff: float, trailing_sl: float) -> float:
    return max(trailing_sl, cur_tick_price - sl_diff)


def trail_sl_in_downward_expected_move(cur_tick_price: float, sl_diff: float, trailing_sl: float) -> float:
    return min(trailing_sl, cur_tick_price + sl_diff)


def sl_hit_in_upward_expected_move(tick_price: float, trailing_sl: float):
    return tick_price <= trailing_sl


def sl_hit_in_downward_expected_move(tick_price: float, trailing_sl: float):
    return tick_price >= trailing_sl


def hit_fixed_point_diff_target_in_up_move(
        cur_tick_price, start_tick_price, fix_point_diff_target) -> bool:
    return cur_tick_price >= start_tick_price + fix_point_diff_target


def hit_fixed_point_diff_target_in_down_move(
        cur_tick_price, start_tick_price, fix_point_diff_target) -> bool:
    return cur_tick_price <= start_tick_price - fix_point_diff_target


def get_market_moves_for_day(
        market_type: MarketType,
        config: Config,
        day: date,
        start_time: time,
        end_time: time,
) -> List[MarketMoveData]:
    price_data: PriceData = fetch_price_data(
        market_type,
        day,
        day,
        start_time,
        end_time,
        config.smooth_price_period,
        config.smooth_price_ema_period,
        config.smooth_slope_period,
        config.smooth_slope_ema_period,
        config.smooth_momentum_period,
        config.smooth_momentum_ema_period,
    )

    price_list: List[PriceDataPerTick] = price_data['price_list']
    n = len(price_list)
    i = 0

    res: List[MarketMoveData] = []

    while i < n - 1:
        price_info: PriceDataPerTick = price_list[i]

        tm: time = price_info['tm']
        if tm < config.min_entry_time:
            i += 1
            continue

        tick_price: float = price_info['tick_price']
        slope: float = price_info['slope']
        smooth_slope: float = price_info['smooth_slope']
        momentum: float = price_info['momentum']
        momentum_rate: float = price_info['momentum_rate']

        if entry_condition_for_up_move(smooth_slope, config.entry_config_smooth_slope_threshold,
                                       momentum, config.entry_config_momentum_threshold,
                                       momentum_rate, config.entry_config_momentum_rate_threshold):
            start_tm: time = tm
            start_tick_price: float = tick_price
            trailing_sl: float = start_tick_price - config.sl

            exit_reason = None
            j = i + 1
            while j < n:
                cur_smooth_price = price_list[j]['smooth_price']
                cur_smooth_price_ema = price_list[j]['smooth_price_ema']
                cur_tick_price = price_list[j]['tick_price']

                # trailing_sl = trail_sl_in_upward_expected_move(cur_tick_price, config.sl, trailing_sl)
                if sl_hit_in_upward_expected_move(cur_tick_price, trailing_sl):
                    exit_reason = exit_reason_sl_hit
                    break

                # if cross_over_happened_in_up_move(cur_smooth_price, cur_smooth_price_ema):
                #     exit_reason = exit_reason_crossover
                #     break

                if hit_fixed_point_diff_target_in_up_move(
                        cur_tick_price, start_tick_price, config.fix_point_diff_target):
                    exit_reason = exit_reason_fix_target_achieved
                    break

                j += 1

            if j == n:
                j -= 1

            end_time: time = price_list[j]['tm']
            end_tick_price: float = price_list[j]['tick_price']
            delta = end_tick_price - start_tick_price

            res.append(MarketMoveData(
                date=day,
                delta=delta,
                expected_direction=direction_up,
                start_time=start_tm,
                end_time=end_time,
                start_point=start_tick_price,
                end_point=end_tick_price,
                exit_reason=exit_reason,
            ))

            i = j + 1
        elif entry_condition_for_down_move(smooth_slope, config.entry_config_smooth_slope_threshold,
                                           momentum, config.entry_config_momentum_threshold,
                                           momentum_rate, config.entry_config_momentum_rate_threshold):
            start_tm: time = tm
            start_tick_price: float = tick_price
            trailing_sl: float = start_tick_price + config.sl

            exit_reason = None
            j = i + 1
            while j < n:
                cur_smooth_price = price_list[j]['smooth_price']
                cur_smooth_price_ema = price_list[j]['smooth_price_ema']
                cur_tick_price = price_list[j]['tick_price']

                # trailing_sl = trail_sl_in_downward_expected_move(cur_tick_price, config.sl, trailing_sl)
                if sl_hit_in_downward_expected_move(cur_tick_price, trailing_sl):
                    exit_reason = exit_reason_sl_hit
                    break

                # if cross_over_happened_in_down_move(cur_smooth_price, cur_smooth_price_ema):
                #     exit_reason = exit_reason_crossover
                #     break

                if hit_fixed_point_diff_target_in_down_move(
                        cur_tick_price, start_tick_price, config.fix_point_diff_target):
                    exit_reason = exit_reason_fix_target_achieved
                    break

                j += 1

            if j == n:
                j -= 1

            end_time: time = price_list[j]['tm']
            end_tick_price: float = price_list[j]['tick_price']
            delta = start_tick_price - price_list[j]['tick_price']

            res.append(MarketMoveData(
                date=day,
                delta=delta,
                expected_direction=direction_down,
                start_time=start_tm,
                end_time=end_time,
                start_point=start_tick_price,
                end_point=end_tick_price,
                exit_reason=exit_reason,
            ))

            i = j + 1
        else:
            i += 1

    return res


def simulate_momentum_analysis(
        config: Config,
        input: Input,
) -> List[MarketMoveData]:
    global_start_date = Input.start_date_time.date()
    global_start_time = Input.start_date_time.time()
    global_end_date = Input.end_date_time.date()
    global_end_time = Input.end_date_time.time()

    res = []

    day = global_start_date
    while day <= global_end_date:
        start_time = global_start_time if day == global_start_date else market_entry_time
        end_time = global_end_time if day == global_end_date else market_exit_time

        print(f'day: {day}, start time: {start_time}, end time: {end_time}')

        res.extend(get_market_moves_for_day(
            input.market_type,
            config,
            day,
            start_time,
            end_time,
        ))

        day += timedelta(days=1)

    return res


def clean_sheet(absolute_file_path, workbook, sheet):
    for row in range(Input.result_start_row, sheet.max_row + 1):
        sheet.cell(row=row, column=Input.col_date).value = None
        sheet.cell(row=row, column=Input.col_entry_time).value = None
        sheet.cell(row=row, column=Input.col_exit_time).value = None
        sheet.cell(row=row, column=Input.col_exp_direction).value = None
        sheet.cell(row=row, column=Input.col_entry_price).value = None
        sheet.cell(row=row, column=Input.col_exit_price).value = None
        sheet.cell(row=row, column=Input.col_move).value = None
        sheet.cell(row=row, column=Input.col_exit_reason).value = None

        sheet.cell(row=row, column=Input.col_momentum_buckets).value = None
        sheet.cell(row=row, column=Input.col_occurrence).value = None
        sheet.cell(row=row, column=Input.col_occurrence_cum_sum).value = None
        sheet.cell(row=row, column=Input.col_occurrence_sum_percentage).value = None

    workbook.save(absolute_file_path)


def write_move_data(absolute_file_path, workbook, sheet, market_moves: List[MarketMoveData]):
    cur_row = Input.result_start_row
    for market_move in market_moves:
        market_move_dict: dict = market_move.to_dict()

        date = market_move_dict['date']
        entry_time = market_move_dict['start_time']
        exit_time = market_move_dict['end_time']
        expected_direction = market_move_dict['expected_direction']
        entry_point = market_move_dict['start_point']
        exit_point = market_move_dict['end_point']
        delta = market_move_dict['delta']
        exit_reason = market_move_dict['exit_reason']

        sheet.cell(row=cur_row, column=Input.col_date, value=date)
        sheet.cell(row=cur_row, column=Input.col_entry_time, value=entry_time)
        sheet.cell(row=cur_row, column=Input.col_exit_time, value=exit_time)
        sheet.cell(row=cur_row, column=Input.col_exp_direction, value=expected_direction)
        sheet.cell(row=cur_row, column=Input.col_entry_price, value=entry_point)
        sheet.cell(row=cur_row, column=Input.col_exit_price, value=exit_point)
        sheet.cell(row=cur_row, column=Input.col_move, value=delta)
        sheet.cell(row=cur_row, column=Input.col_exit_reason, value=exit_reason)

        cur_row += 1

    workbook.save(absolute_file_path)


def write_momentum_occurrences(absolute_file_path, workbook, sheet, momentum_distribution: dict):
    cur_row = Input.result_start_row
    for group_range, arr in momentum_distribution.items():
        count = arr[0]
        cum_sum = arr[1]
        cum_percentage = arr[2]

        sheet.cell(row=cur_row, column=Input.col_momentum_buckets, value=group_range)
        sheet.cell(row=cur_row, column=Input.col_occurrence, value=count)
        sheet.cell(row=cur_row, column=Input.col_occurrence_cum_sum, value=cum_sum)
        sheet.cell(row=cur_row, column=Input.col_occurrence_sum_percentage, value=cum_percentage)

        cur_row += 1

    workbook.save(absolute_file_path)


def save_in_output_xlsx_sheet(market_moves: List[MarketMoveData], momentum_distribution: dict):
    absolute_file_path = os.path.abspath(Input.output_file_relative_path)
    workbook = openpyxl.load_workbook(absolute_file_path)
    sheet = workbook[Input.sheet_name]

    if Input.clean_sheet:
        clean_sheet(absolute_file_path, workbook, sheet)

    write_move_data(absolute_file_path, workbook, sheet, market_moves)
    write_momentum_occurrences(absolute_file_path, workbook, sheet, momentum_distribution)

    workbook.close()


def main():
    market_moves: List[MarketMoveData] = simulate_momentum_analysis(Config(), Input())

    # print(f'market_moves: {market_moves}')

    move_deltas = [market_move.delta for market_move in market_moves]
    momentum_distribution: dict = get_occurrence_distribution(
        move_deltas,
        Input.momentum_occurrence_distribution_bucket_size,
    )

    # print(f'momentum_distribution: {momentum_distribution}')

    save_in_output_xlsx_sheet(market_moves, momentum_distribution)
